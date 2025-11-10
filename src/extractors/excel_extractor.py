"""
Excel Extractor - Microsoft Excel Workbook Extraction

Extracts content from Excel workbooks (.xlsx, .xls) with support for:
- Multi-sheet workbooks
- Cell values and formulas
- Table structure preservation
- Chart metadata extraction
- Cross-sheet references

Implementation follows strict TDD methodology with infrastructure integration.
"""

from pathlib import Path
from typing import Optional, Union
from datetime import datetime, timezone
import hashlib
import logging
import time
import warnings as warnings_module

try:
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException
except ImportError:
    raise ImportError(
        "openpyxl is required for ExcelExtractor. " "Install with: pip install openpyxl"
    )

import sys

sys.path.insert(0, str(Path(__file__).parent.parent))

from core import (
    BaseExtractor,
    ContentBlock,
    ContentType,
    DocumentMetadata,
    ExtractionResult,
    Position,
    TableMetadata,
)

# Import infrastructure components
try:
    from infrastructure import (
        ConfigManager,
        get_logger,
        ErrorHandler,
        ProgressTracker,
        RecoveryAction,
    )

    INFRASTRUCTURE_AVAILABLE = True
except ImportError:
    INFRASTRUCTURE_AVAILABLE = False


class ExcelExtractor(BaseExtractor):
    """
    Extracts content from Microsoft Excel workbooks.

    Supports multi-sheet extraction with cell values, formulas, and structure
    preservation. Uses openpyxl for Excel file parsing.

    Design Notes:
    - Extracts all sheets by default
    - Each sheet becomes a TABLE ContentBlock
    - Cell data stored in TableMetadata
    - Formulas preserved in metadata
    - Charts detected and metadata extracted

    Example:
        >>> extractor = ExcelExtractor()
        >>> result = extractor.extract(Path("workbook.xlsx"))
        >>> if result.success:
        ...     for block in result.content_blocks:
        ...         print(f"Sheet: {block.position.sheet}")
    """

    def __init__(self, config: Optional[Union[dict, object]] = None):
        """
        Initialize Excel extractor with optional configuration.

        Args:
            config: Configuration options (dict or ConfigManager):
                - max_rows: Maximum rows to extract per sheet (default: None)
                - max_columns: Maximum columns per sheet (default: None)
                - include_formulas: Extract formula strings (default: True)
                - include_charts: Extract chart metadata (default: True)
                - skip_empty_cells: Skip empty cells (default: False)
        """
        super().__init__(config if isinstance(config, dict) or config is None else {})

        # Detect ConfigManager
        is_config_manager = (
            INFRASTRUCTURE_AVAILABLE
            and hasattr(config, "__class__")
            and config.__class__.__name__ == "ConfigManager"
        )
        self._config_manager = config if is_config_manager else None

        # Initialize infrastructure
        if INFRASTRUCTURE_AVAILABLE:
            self.logger = get_logger(__name__)
            self.error_handler = ErrorHandler()
        else:
            self.logger = logging.getLogger(__name__)
            self.error_handler = None

        # Load configuration
        if self._config_manager:
            # From ConfigManager
            extractor_config = self._config_manager.get_section("extractors.excel", default={})
            self.max_rows = extractor_config.get("max_rows")
            self.max_columns = extractor_config.get("max_columns")
            include_formulas_val = extractor_config.get("include_formulas")
            self.include_formulas = (
                include_formulas_val if include_formulas_val is not None else True
            )
            include_charts_val = extractor_config.get("include_charts")
            self.include_charts = include_charts_val if include_charts_val is not None else True
            skip_empty_val = extractor_config.get("skip_empty_cells")
            self.skip_empty_cells = skip_empty_val if skip_empty_val is not None else False
        elif isinstance(config, dict):
            # From dict (backward compatible)
            self.max_rows = config.get("max_rows", None)
            self.max_columns = config.get("max_columns", None)
            self.include_formulas = config.get("include_formulas", True)
            self.include_charts = config.get("include_charts", True)
            self.skip_empty_cells = config.get("skip_empty_cells", False)
        else:
            # Defaults
            self.max_rows = None
            self.max_columns = None
            self.include_formulas = True
            self.include_charts = True
            self.skip_empty_cells = False

    def supports_format(self, file_path: Path) -> bool:
        """
        Check if file is an Excel file.

        Args:
            file_path: Path to check

        Returns:
            True if file has .xlsx or .xls extension
        """
        return file_path.suffix.lower() in [".xlsx", ".xls"]

    def get_supported_extensions(self) -> list[str]:
        """Return supported file extensions."""
        return [".xlsx", ".xls"]

    def get_format_name(self) -> str:
        """Return human-readable format name."""
        return "Microsoft Excel"

    def extract(self, file_path: Path) -> ExtractionResult:
        """
        Extract content from Excel workbook.

        Strategy:
        1. Validate file exists and is accessible
        2. Open workbook with openpyxl
        3. Extract each sheet sequentially
        4. Create TableMetadata for each sheet
        5. Extract cell values and formulas
        6. Detect charts if enabled
        7. Generate document metadata
        8. Return structured result

        Args:
            file_path: Path to Excel file

        Returns:
            ExtractionResult with content blocks and metadata
        """
        start_time = time.time()

        # Log extraction start
        if INFRASTRUCTURE_AVAILABLE:
            self.logger.info("Starting Excel extraction", extra={"file": str(file_path)})

        errors = []
        warnings = []
        content_blocks = []
        tables = []

        # Step 1: Validate file
        is_valid, validation_errors = self.validate_file(file_path)
        if not is_valid:
            if self.error_handler:
                error = self.error_handler.create_error("E001", file_path=str(file_path))
                errors.append(self.error_handler.format_for_user(error))
                if INFRASTRUCTURE_AVAILABLE:
                    self.logger.error("File validation failed", extra={"file": str(file_path)})
            else:
                errors.extend(validation_errors)

            return ExtractionResult(
                success=False,
                errors=tuple(errors),
                document_metadata=DocumentMetadata(
                    source_file=file_path,
                    file_format="xlsx",
                ),
            )

        try:
            # Step 2: Open workbook
            try:
                # Suppress openpyxl UserWarning about workbook default style
                # This warning appears for some Excel files but doesn't affect extraction
                with warnings_module.catch_warnings():
                    warnings_module.filterwarnings(
                        "ignore", message="Workbook contains no default style"
                    )

                    # Load with data_only=False to get formulas, then reload with data_only=True for values
                    wb = load_workbook(file_path, data_only=False)
                    # Keep reference to formula workbook
                    wb_values = None
                    if self.include_formulas:
                        try:
                            wb_values = load_workbook(file_path, data_only=True)
                        except:
                            pass  # If we can't load values, use formulas only
            except InvalidFileException as e:
                if self.error_handler:
                    error = self.error_handler.create_error(
                        "E171", file_path=str(file_path), original_exception=e
                    )
                    errors.append(self.error_handler.format_for_user(error))
                else:
                    errors.append(f"Invalid Excel file format: {str(e)}")

                return ExtractionResult(
                    success=False,
                    errors=tuple(errors),
                    document_metadata=DocumentMetadata(
                        source_file=file_path,
                        file_format="xlsx",
                    ),
                )
            except Exception as e:
                return ExtractionResult(
                    success=False,
                    errors=(f"Failed to open Excel file: {str(e)}",),
                    document_metadata=DocumentMetadata(
                        source_file=file_path,
                        file_format="xlsx",
                    ),
                )

            # Step 3: Extract all sheets
            sheet_count = 0
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ws_values = wb_values[sheet_name] if wb_values else None

                # Extract sheet content
                sheet_blocks, sheet_table = self._extract_sheet(
                    ws, sheet_name, sheet_count, ws_values
                )

                content_blocks.extend(sheet_blocks)
                if sheet_table:
                    tables.append(sheet_table)

                sheet_count += 1

            # Step 4: Generate document metadata
            doc_metadata = self._extract_document_metadata(file_path, wb)

            # Update statistics
            doc_metadata = DocumentMetadata(
                source_file=doc_metadata.source_file,
                file_format=doc_metadata.file_format,
                file_size_bytes=doc_metadata.file_size_bytes,
                file_hash=doc_metadata.file_hash,
                title=doc_metadata.title,
                author=doc_metadata.author,
                created_date=doc_metadata.created_date,
                modified_date=doc_metadata.modified_date,
                subject=doc_metadata.subject,
                keywords=doc_metadata.keywords,
                table_count=len(tables),
                extracted_at=doc_metadata.extracted_at,
                extractor_version="0.1.0-tdd",
            )

            # Step 5: Log completion and return result
            duration = time.time() - start_time
            if INFRASTRUCTURE_AVAILABLE:
                self.logger.info(
                    "Excel extraction complete",
                    extra={
                        "file": str(file_path),
                        "sheets": sheet_count,
                        "tables": len(tables),
                        "blocks": len(content_blocks),
                        "duration_seconds": round(duration, 3),
                    },
                )

            return ExtractionResult(
                content_blocks=tuple(content_blocks),
                document_metadata=doc_metadata,
                tables=tuple(tables),
                success=True,
                warnings=tuple(warnings),
            )

        except PermissionError as e:
            if self.error_handler:
                error = self.error_handler.create_error("E500", file_path=str(file_path))
                errors.append(self.error_handler.format_for_user(error))
            else:
                errors.append(f"Permission denied reading file: {file_path}")

            if INFRASTRUCTURE_AVAILABLE:
                self.logger.error("Permission denied", extra={"file": str(file_path)})

        except Exception as e:
            if self.error_handler:
                error = self.error_handler.create_error(
                    "E170", file_path=str(file_path), original_exception=e
                )
                errors.append(self.error_handler.format_for_user(error))
            else:
                errors.append(f"Unexpected error during extraction: {str(e)}")

            if INFRASTRUCTURE_AVAILABLE:
                self.logger.error(
                    "Extraction failed", extra={"file": str(file_path), "error": str(e)}
                )

        # Return failed result
        return ExtractionResult(
            success=False,
            errors=tuple(errors),
            warnings=tuple(warnings),
            document_metadata=DocumentMetadata(
                source_file=file_path,
                file_format="xlsx",
            ),
        )

    def _extract_sheet(
        self, worksheet, sheet_name: str, sheet_index: int, worksheet_values=None
    ) -> tuple[list[ContentBlock], Optional[TableMetadata]]:
        """
        Extract content from a single worksheet.

        Args:
            worksheet: openpyxl Worksheet object (with formulas)
            sheet_name: Name of the sheet
            sheet_index: Index of sheet in workbook
            worksheet_values: Optional worksheet with calculated values

        Returns:
            Tuple of (content_blocks, table_metadata)
        """
        content_blocks = []

        # Get sheet dimensions
        max_row = worksheet.max_row
        max_col = worksheet.max_column

        # Apply limits if configured
        if self.max_rows:
            max_row = min(max_row, self.max_rows)
        if self.max_columns:
            max_col = min(max_col, self.max_columns)

        # Extract cell data
        cells_data = []
        formulas = {}
        has_formulas = False

        for row_idx in range(1, max_row + 1):
            row_data = []
            for col_idx in range(1, max_col + 1):
                cell = worksheet.cell(row_idx, col_idx)

                # Get cell value - prefer calculated value if available
                if worksheet_values:
                    cell_value = worksheet_values.cell(row_idx, col_idx).value
                else:
                    cell_value = cell.value

                if cell_value is None:
                    cell_value = ""

                row_data.append(str(cell_value))

                # Extract formula if present and enabled
                if self.include_formulas and hasattr(cell, "data_type") and cell.data_type == "f":
                    try:
                        formula_value = (
                            cell.value
                            if isinstance(cell.value, str) and cell.value.startswith("=")
                            else None
                        )
                        if formula_value:
                            cell_ref = f"{chr(64 + col_idx)}{row_idx}"
                            formulas[cell_ref] = formula_value
                            has_formulas = True
                    except:
                        pass  # Ignore formula extraction errors

            cells_data.append(tuple(row_data))

        # Create TableMetadata
        table_metadata = TableMetadata(
            num_rows=max_row,
            num_columns=max_col,
            has_header=max_row > 0,  # Assume first row is header if data exists
            header_row=tuple(cells_data[0]) if cells_data else None,
            cells=tuple(cells_data),
        )

        # Create ContentBlock for sheet
        block_metadata = {
            "sheet_name": sheet_name,
            "sheet_index": sheet_index,
            "num_rows": max_row,
            "num_columns": max_col,
            "table_id": table_metadata.table_id,
        }

        if has_formulas:
            block_metadata["has_formulas"] = True
            block_metadata["formulas"] = formulas

        # Create TABLE content block
        block = ContentBlock(
            block_type=ContentType.TABLE,
            content=f"Sheet: {sheet_name}",
            position=Position(sheet=sheet_name, sequence_index=sheet_index),
            confidence=1.0,
            metadata=block_metadata,
        )

        content_blocks.append(block)

        return content_blocks, table_metadata

    def _extract_document_metadata(self, file_path: Path, workbook) -> DocumentMetadata:
        """
        Extract document-level metadata from Excel file.

        Args:
            file_path: Path to file
            workbook: openpyxl Workbook object

        Returns:
            DocumentMetadata with available properties
        """
        # File system metadata
        file_stat = file_path.stat()
        file_size = file_stat.st_size

        # Generate file hash
        file_hash = self._compute_file_hash(file_path)

        # Extract workbook properties
        props = workbook.properties

        # Parse dates
        created = props.created
        modified = props.modified

        # Parse keywords
        keywords = ()
        if props.keywords:
            keywords = tuple(k.strip() for k in props.keywords.split(",") if k.strip())

        return DocumentMetadata(
            source_file=file_path,
            file_format="xlsx",
            file_size_bytes=file_size,
            file_hash=file_hash,
            title=props.title or None,
            author=props.creator or None,
            created_date=created if isinstance(created, datetime) else None,
            modified_date=modified if isinstance(modified, datetime) else None,
            subject=props.subject or None,
            keywords=keywords,
            extracted_at=datetime.now(timezone.utc),
        )

    def _compute_file_hash(self, file_path: Path) -> str:
        """
        Compute SHA256 hash of file for deduplication.

        Args:
            file_path: Path to file

        Returns:
            Hex string of SHA256 hash
        """
        sha256 = hashlib.sha256()

        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(8192), b""):
                sha256.update(chunk)

        return sha256.hexdigest()
