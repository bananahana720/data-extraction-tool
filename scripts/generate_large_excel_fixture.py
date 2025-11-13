"""
Generate large Excel fixture for integration testing.

Creates a 10K+ row synthetic audit data spreadsheet.
Uses openpyxl to generate Excel file with realistic audit data structure.
"""

import random
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def generate_large_audit_excel(output_path: Path, num_rows: int = 10240) -> None:
    """Generate a large (10K+ rows) synthetic audit data Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Data"

    # Header row styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # Define headers
    headers = [
        "Risk ID",
        "Risk Description",
        "Impact",
        "Likelihood",
        "Risk Rating",
        "Control ID",
        "Control Description",
        "Control Owner",
        "Control Status",
        "Test Date",
        "Test Result",
        "Findings",
        "Action Required",
        "Target Date",
    ]

    # Write headers
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Data generation templates
    risk_descriptions = [
        "Market volatility impact on revenue streams",
        "Supply chain disruption affecting operations",
        "Talent retention and recruitment challenges",
        "Regulatory compliance requirements gaps",
        "Cybersecurity threat exposure vulnerabilities",
        "Data privacy and protection concerns",
        "Operational process inefficiencies",
        "Third-party vendor management risks",
        "Financial reporting accuracy issues",
        "Technology infrastructure dependencies",
        "Strategic planning execution gaps",
        "Customer satisfaction and retention",
        "Product quality and safety standards",
        "Environmental sustainability requirements",
        "Business continuity planning deficiencies",
    ]

    control_descriptions = [
        "Segregation of duties verification process",
        "Authorization approval workflow controls",
        "Monthly reconciliation procedures",
        "Quarterly access control reviews",
        "Change management approval process",
        "Data backup and recovery testing",
        "Vendor due diligence assessments",
        "Financial close process controls",
        "IT security monitoring procedures",
        "Compliance training program delivery",
        "Incident response plan testing",
        "Quality assurance inspections",
        "Performance monitoring dashboards",
        "Risk assessment annual reviews",
        "Policy acknowledgment tracking",
    ]

    control_owners = [
        "Finance Department",
        "IT Security Team",
        "Operations Manager",
        "Compliance Officer",
        "Risk Management",
        "Internal Audit",
        "HR Department",
        "Legal Department",
        "Supply Chain",
        "Quality Assurance",
    ]

    impacts = ["Low", "Medium", "High", "Critical"]
    likelihoods = ["Low", "Medium", "High"]
    risk_ratings = ["Low", "Medium", "High", "Critical"]
    control_statuses = ["Effective", "Partially Effective", "Ineffective", "Not Tested"]
    test_results = ["Pass", "Pass with Exceptions", "Fail"]
    findings_options = ["None", "1 Minor", "2 Minor", "1 Major", "Multiple Issues"]
    action_required = ["None", "Remediation Plan", "Immediate Action", "Monitoring"]

    # Generate rows
    start_date = date(2024, 1, 1)

    print(f"Generating {num_rows - 1} rows of audit data...")

    for row_num in range(2, num_rows + 1):
        risk_id = f"R-{row_num - 1:05d}"
        risk_desc = random.choice(risk_descriptions)
        impact = random.choice(impacts)
        likelihood = random.choice(likelihoods)

        # Calculate risk rating based on impact and likelihood
        if impact == "Critical" or (impact == "High" and likelihood == "High"):
            risk_rating = "Critical"
        elif impact == "High" or likelihood == "High":
            risk_rating = "High"
        elif impact == "Medium" and likelihood == "Medium":
            risk_rating = "Medium"
        else:
            risk_rating = "Low"

        control_id = f"C-{row_num - 1:05d}"
        control_desc = random.choice(control_descriptions)
        control_owner = random.choice(control_owners)
        control_status = random.choice(control_statuses)

        # Test date within last year
        days_ago = random.randint(0, 365)
        test_date = start_date + timedelta(days=days_ago)

        test_result = random.choice(test_results)
        findings = random.choice(findings_options)
        action = random.choice(action_required)

        # Target date 90-120 days from test date
        target_days = random.randint(90, 120)
        target_date = test_date + timedelta(days=target_days)

        # Write row data
        row_data = [
            risk_id,
            risk_desc,
            impact,
            likelihood,
            risk_rating,
            control_id,
            control_desc,
            control_owner,
            control_status,
            test_date,
            test_result,
            findings,
            action,
            target_date,
        ]

        for col_num, value in enumerate(row_data, start=1):
            ws.cell(row=row_num, column=col_num, value=value)

        # Progress indicator
        if row_num % 1000 == 0:
            print(f"  Generated {row_num - 1} rows...")

    # Adjust column widths
    column_widths = [12, 40, 10, 12, 12, 12, 40, 18, 18, 12, 18, 18, 18, 12]
    for col_num, width in enumerate(column_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save workbook
    wb.save(output_path)

    print(f"Generated large Excel: {output_path}")
    print(f"Total rows: {num_rows} (including header)")
    print(f"File size: {output_path.stat().st_size / (1024 * 1024):.2f} MB")


if __name__ == "__main__":
    output_dir = Path(__file__).parent.parent / "tests" / "fixtures" / "xlsx" / "large"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_file = output_dir / "audit-data-10k-rows.xlsx"
    generate_large_audit_excel(output_file, num_rows=10240)
